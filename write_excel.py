from openpyxl import Workbook 

wb = Workbook()
ws = wb.active

ws["A1"] = "Hello"
ws["B2"] = "World"
ws["C3"] = "from"
ws["D4"] = "Python"

wb.save("sample2.xlsx")
